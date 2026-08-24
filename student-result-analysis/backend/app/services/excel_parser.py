from __future__ import annotations

import re
from dataclasses import dataclass, field
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

MAX_FILE_BYTES = 10 * 1024 * 1024
USN_RE = re.compile(r"^[0-9][A-Z0-9]{7,19}$")
SUBJECT_CODE_RE = re.compile(r"^[A-Z]{1,5}\d{2}[A-Z]{2,5}\d{2,3}$")
ABSENT_MARKS = {"AB", "A", "NE", "MP", "AA", "ABSENT", "-"}


@dataclass
class ImportErrorItem:
    row: int
    usn: str | None
    subject: str | None
    error: str


@dataclass
class SubjectColumns:
    code: str
    ia_col: int
    ext_col: int
    total_col: int


@dataclass
class ParsedMark:
    subject_code: str
    internal_marks: float
    external_marks: float
    total_marks: float


@dataclass
class ParsedStudent:
    row: int
    slno: int | None
    usn: str
    name: str
    marks: list[ParsedMark]


@dataclass
class ParsedWorkbook:
    sheet_name: str
    department: str
    semester: int
    academic_year: str | None
    subjects: list[SubjectColumns]
    students: list[ParsedStudent]
    errors: list[ImportErrorItem] = field(default_factory=list)


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _cell_upper(value: Any) -> str:
    return _cell_str(value).upper()


def _unmerge(ws: Worksheet) -> list[list[Any]]:
    # Copy merged values only across the top row of each merge. Filling the
    # whole range would overwrite IA/Ext/T labels under merged subject codes.
    filled: dict[tuple[int, int], Any] = {}
    for merged in list(ws.merged_cells.ranges):
        value = ws.cell(merged.min_row, merged.min_col).value
        for col in range(merged.min_col, merged.max_col + 1):
            filled[(merged.min_row, col)] = value

    grid: list[list[Any]] = []
    for r in range(1, (ws.max_row or 1) + 1):
        row_vals: list[Any] = []
        for c in range(1, (ws.max_column or 1) + 1):
            if (r, c) in filled:
                row_vals.append(filled[(r, c)])
            else:
                row_vals.append(ws.cell(r, c).value)
        grid.append(row_vals)
    return grid


def _at(grid: list[list[Any]], row_idx: int, col_idx: int) -> Any:
    if row_idx < 0 or row_idx >= len(grid):
        return None
    row = grid[row_idx]
    if col_idx < 0 or col_idx >= len(row):
        return None
    return row[col_idx]


def _parse_semester(text: str) -> int | None:
    match = re.search(r"(\d+)\s*(?:ST|ND|RD|TH)?\s*SEMESTER", text, re.I)
    if match:
        value = int(match.group(1))
        if 1 <= value <= 8:
            return value
    match = re.search(r"\bSEM(?:ESTER)?\s*[-:]?\s*(\d+)\b", text, re.I)
    if match:
        value = int(match.group(1))
        if 1 <= value <= 8:
            return value
    roman = {"I": 1, "II": 2, "III": 3, "IV": 4, "V": 5, "VI": 6, "VII": 7, "VIII": 8}
    match = re.search(r"\b(VIII|VII|VI|IV|V|III|II|I)\s*(?:ST)?\s*SEM", text, re.I)
    if match:
        return roman[match.group(1).upper()]
    return None


def _parse_department(text: str) -> str:
    upper = text.upper()
    if "MASTER OF COMPUTER" in upper or re.search(r"\bMCA\b", upper):
        return "MCA"
    match = re.search(r"DEPARTMENT OF\s+([A-Z &]+)", upper)
    if match:
        return match.group(1).strip().title()[:80]
    return "MCA"


def _parse_academic_year(text: str) -> str | None:
    match = re.search(r"AY\s*(\d{4}\s*[-–]\s*\d{2,4})", text, re.I)
    if match:
        return re.sub(r"\s+", "", match.group(1).replace("–", "-"))
    match = re.search(r"(20\d{2})\s*[-–]\s*(\d{2,4})", text)
    if match:
        return f"{match.group(1)}-{match.group(2)}"
    return None


def _to_mark(value: Any, *, row: int, usn: str, subject: str, field: str, errors: list[ImportErrorItem]) -> float | None:
    raw = _cell_str(value)
    if value is None or raw == "":
        errors.append(ImportErrorItem(row, usn, subject, f"Missing {field} marks"))
        return None
    if raw.upper() in ABSENT_MARKS:
        return 0.0
    if raw.startswith("="):
        errors.append(ImportErrorItem(row, usn, subject, f"{field} is a formula without a cached value"))
        return None
    try:
        number = float(raw.replace("%", ""))
    except (TypeError, ValueError):
        errors.append(ImportErrorItem(row, usn, subject, f"Invalid {field} marks: {raw}"))
        return None
    if number < 0 or number > 100:
        errors.append(ImportErrorItem(row, usn, subject, f"{field} marks must be between 0 and 100"))
        return None
    return number


def _find_header_row(grid: list[list[Any]]) -> int:
    for idx, row in enumerate(grid):
        labels = {_cell_upper(v).replace(".", "").replace(" ", "") for v in row}
        if "USN" in {_cell_upper(v) for v in row} and any(
            "NAME" in _cell_upper(v) for v in row
        ):
            return idx
    raise ValueError("Could not find the USN / STUDENT NAME header row")


def _name_col(row: list[Any]) -> int | None:
    for idx, value in enumerate(row):
        label = _cell_upper(value)
        if "STUDENT" in label and "NAME" in label:
            return idx
    for idx, value in enumerate(row):
        if _cell_upper(value) in {"NAME", "STUDENT NAME", "STUDENTNAME"}:
            return idx
    return None


def _looks_like_code(value: Any) -> bool:
    text = re.sub(r"\s+", "", _cell_upper(value))
    if SUBJECT_CODE_RE.match(text):
        return True
    return bool(
        re.match(r"^[A-Z0-9]{6,16}$", text)
        and any(ch.isdigit() for ch in text)
        and any(ch.isalpha() for ch in text)
        and not _is_summary_col(text)
    )


def _normalize_code(value: Any) -> str:
    return re.sub(r"\s+", "", _cell_upper(value))


def _find_col(row: list[Any], *names: str) -> int | None:
    wanted = {n.upper() for n in names}
    for idx, value in enumerate(row):
        if _cell_upper(value) in wanted:
            return idx
    return None


def _is_ia(value: Any) -> bool:
    return _cell_upper(value) in {"IA", "CIA", "CIE", "INTERNAL"}


def _is_ext(value: Any) -> bool:
    return _cell_upper(value) in {"EXT", "SEE", "EXTERNAL", "EX"}


def _is_total_sub(value: Any) -> bool:
    return _cell_upper(value) in {"T", "TOT", "TOTAL"}


def _is_summary_col(value: Any) -> bool:
    return _cell_upper(value) in {"TOTAL", "TOT", "AVG", "AVERAGE", "AV"}


def _subject_groups(header: list[Any], sub: list[Any], above: list[Any], start_col: int) -> list[SubjectColumns]:
    subjects: list[SubjectColumns] = []
    width = max(len(header), len(sub), len(above))
    col = start_col

    def code_at(c: int) -> str:
        for row in (header, above, sub):
            if c < len(row) and _looks_like_code(row[c]) and not _is_summary_col(row[c]):
                return _normalize_code(row[c])
        return ""

    while col < width:
        h = _cell_upper(header[col]) if col < len(header) else ""
        s = _cell_upper(sub[col]) if col < len(sub) else ""
        if _is_summary_col(h) or _is_summary_col(s):
            col += 1
            continue

        ia = ext = tot = None
        code = ""
        for offset in range(3):
            c = col + offset
            if c >= width:
                break
            label = _cell_upper(sub[c]) if c < len(sub) else ""
            code = code or code_at(c)
            if _is_ia(label):
                ia = c
            elif _is_ext(label):
                ext = c
            elif label in {"T", "TOT"}:
                tot = c
        if ia is not None and ext is not None and tot is not None and code:
            subjects.append(SubjectColumns(code=code, ia_col=ia, ext_col=ext, total_col=tot))
            col = max(ia, ext, tot) + 1
            continue
        col += 1

    if subjects:
        return subjects

    col = start_col
    while col + 2 < width:
        h = _cell_upper(header[col]) if col < len(header) else ""
        s = _cell_upper(sub[col]) if col < len(sub) else ""
        if _is_summary_col(h) or _is_summary_col(s):
            break
        code = code_at(col) or code_at(col + 1) or code_at(col + 2)
        if not code:
            col += 1
            continue
        subjects.append(SubjectColumns(code=code, ia_col=col, ext_col=col + 1, total_col=col + 2))
        col += 3
    return subjects


def _pick_sheet(workbook) -> Any:
    named = [
        name
        for name in workbook.sheetnames
        if name.strip().lower().replace("_", " ") in {"data entry", "dataentry", "result"}
    ]
    if named:
        return workbook[named[0]]
    for name in workbook.sheetnames:
        ws = workbook[name]
        grid = _unmerge(ws)
        for row in grid[:40]:
            if "USN" in {_cell_upper(v) for v in row}:
                return ws
    return workbook[workbook.sheetnames[0]]


def parse_result_workbook(content: bytes, filename: str) -> ParsedWorkbook:
    if len(content) > MAX_FILE_BYTES:
        raise ValueError("File is larger than 10MB")
    lower = filename.lower()
    if not (lower.endswith(".xlsx") or lower.endswith(".xlsm")):
        raise ValueError("Upload an Excel workbook (.xlsx) in the Result sheet Sample format")

    try:
        workbook = load_workbook(BytesIO(content), data_only=False)
    except Exception as exc:  # noqa: BLE001
        raise ValueError("Could not read the Excel file. Confirm it is a valid .xlsx workbook.") from exc

    ws = _pick_sheet(workbook)
    grid = _unmerge(ws)
    if not grid:
        raise ValueError("The Excel sheet is empty")

    banner = " ".join(_cell_str(v) for row in grid[:25] for v in row)
    department = _parse_department(banner)
    semester = _parse_semester(banner)
    if semester is None:
        raise ValueError(
            "Could not read the semester from the sheet title (expected e.g. 1st SEMESTER). "
            f"Using sheet '{ws.title}'."
        )
    academic_year = _parse_academic_year(banner)

    header_idx = _find_header_row(grid)
    header = grid[header_idx]
    sub = grid[header_idx + 1] if header_idx + 1 < len(grid) else []
    above = grid[header_idx - 1] if header_idx > 0 else []

    usn_col = _find_col(header, "USN")
    name_col = _name_col(header)
    slno_col = _find_col(header, "SL.NO", "SL NO", "SLNO", "S.NO", "SI.NO", "SL. NO")
    if usn_col is None or name_col is None:
        raise ValueError("Missing required columns USN and STUDENT NAME")

    start_col = max(usn_col, name_col) + 1
    subjects = _subject_groups(header, sub, above, start_col)
    if not subjects:
        raise ValueError(
            "Could not find subject columns. Expected subject codes with IA / Ext / T groups "
            f"on sheet '{ws.title}'."
        )

    errors: list[ImportErrorItem] = []
    students: list[ParsedStudent] = []
    seen_usn: dict[str, int] = {}

    for row_idx in range(header_idx + 1, len(grid)):
        excel_row = row_idx + 1
        usn = _cell_upper(_at(grid, row_idx, usn_col)).replace(" ", "")
        name = _cell_str(_at(grid, row_idx, name_col))
        if not usn and not name:
            continue
        if not USN_RE.match(usn):
            continue
        if not name:
            errors.append(ImportErrorItem(excel_row, usn, None, "Missing student name"))
            continue
        if usn in seen_usn:
            errors.append(
                ImportErrorItem(excel_row, usn, None, f"Duplicate USN in file (also on row {seen_usn[usn]})")
            )
            continue
        seen_usn[usn] = excel_row

        slno = None
        if slno_col is not None:
            raw = _at(grid, row_idx, slno_col)
            if raw not in (None, ""):
                try:
                    slno = int(float(raw))
                except (TypeError, ValueError):
                    slno = None

        marks: list[ParsedMark] = []
        for subject in subjects:
            ia = _to_mark(_at(grid, row_idx, subject.ia_col), row=excel_row, usn=usn, subject=subject.code, field="IA", errors=errors)
            ext = _to_mark(_at(grid, row_idx, subject.ext_col), row=excel_row, usn=usn, subject=subject.code, field="Ext", errors=errors)
            if ia is None or ext is None:
                continue
            computed = round(ia + ext, 2)
            if computed > 100:
                errors.append(ImportErrorItem(excel_row, usn, subject.code, "IA+Ext exceeds 100"))
                continue
            marks.append(
                ParsedMark(
                    subject_code=subject.code,
                    internal_marks=ia,
                    external_marks=ext,
                    total_marks=computed,
                )
            )

        if len(marks) != len(subjects):
            if not any(e.row == excel_row for e in errors):
                errors.append(ImportErrorItem(excel_row, usn, None, "Incomplete subject marks"))
            continue

        students.append(ParsedStudent(row=excel_row, slno=slno, usn=usn, name=name, marks=marks))

    if not students and not errors:
        raise ValueError(f"No student rows found under the header on sheet '{ws.title}'")

    return ParsedWorkbook(
        sheet_name=ws.title,
        department=department,
        semester=semester,
        academic_year=academic_year,
        subjects=subjects,
        students=students,
        errors=errors,
    )
